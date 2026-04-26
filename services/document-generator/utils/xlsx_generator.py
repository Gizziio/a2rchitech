from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import tempfile


def create_rubric_spreadsheet(
    template_type: str,
    title: Optional[str],
    columns: List[str],
    rows: List[List[str]],
    formulas: Optional[Dict[str, str]],
    citations: Optional[List[dict]]
) -> str:
    """
    Create a rubric spreadsheet in XLSX format.
    
    Args:
        template_type: Type of template (rubric, tracker, study_plan, grade_calc)
        title: Optional title for the spreadsheet
        columns: Column headers
        rows: Data rows
        formulas: Optional formulas to apply
        citations: Optional citations to include
        
    Returns:
        Path to the generated XLSX file
    """
    wb = Workbook()
    ws = wb.active
    
    if title:
        ws.title = title[:31]  # Excel sheet names limited to 31 chars
    else:
        ws.title = template_type.title()[:31]
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add title row if provided
    if title:
        ws['A1'] = title
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:E1')  # Adjust range based on number of columns
        ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add column headers
    start_row = 2 if title else 1
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Auto-adjust column width
        ws.column_dimensions[get_column_letter(col_idx)].width = len(column) + 5
    
    # Add data rows
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply border to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Add formulas if provided
    if formulas:
        # Add a formulas section after the data
        formulas_start_row = len(rows) + start_row + 2
        ws.cell(row=formulas_start_row, column=1, value="Formulas").font = Font(bold=True)
        
        formula_row = formulas_start_row + 1
        for formula_name, formula_value in formulas.items():
            ws.cell(row=formula_row, column=1, value=formula_name)
            ws.cell(row=formula_row, column=2, value=formula_value)
            formula_row += 1
    
    # Add citations sheet if provided
    if citations:
        citations_ws = wb.create_sheet("References")
        citations_ws['A1'] = 'References'
        citations_ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        
        row = 3
        for citation in citations:
            citations_ws.cell(row=row, column=1, value=citation['title'])
            citations_ws.cell(row=row, column=2, value=citation['url'])
            row += 1
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name